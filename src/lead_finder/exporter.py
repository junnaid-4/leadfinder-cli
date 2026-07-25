"""Export businesses and their analysis data to structured files."""

from __future__ import annotations

import csv
import json
import os
import sqlite3
import tempfile
from dataclasses import dataclass
from pathlib import Path
from typing import Any, cast

import openpyxl
from openpyxl.cell.cell import Cell
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.properties import PageSetupProperties
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.worksheet import Worksheet


@dataclass(frozen=True)
class LeadExportRow:
    """A flattened representation of a business and its latest checks."""

    business_id: int
    place_id: str
    business_name: str
    phone: str | None
    address: str | None
    rating: float | None
    review_count: int | None
    business_status: str | None
    original_website: str | None
    normalized_website: str | None
    final_website: str | None
    website_status: str | None
    http_status: int | None
    website_checked_at: str | None
    raw_score: int | None
    final_score: int | None
    priority: str | None
    scoring_version: str | None
    score_breakdown: str | None
    scored_at: str | None
    discovery_queries: str | None
    google_maps_url: str | None

    @classmethod
    def from_sqlite_row(cls, row: sqlite3.Row) -> LeadExportRow:
        """Create from a database row."""
        return cls(
            business_id=row["business_id"],
            place_id=row["place_id"],
            business_name=row["business_name"],
            phone=row["phone"],
            address=row["address"],
            rating=row["rating"],
            review_count=row["review_count"],
            business_status=row["business_status"],
            original_website=row["website_original_url"],
            normalized_website=row["website_normalized_url"],
            final_website=row["website_final_url"],
            website_status=row["website_status"],
            http_status=row["http_status"],
            website_checked_at=row["website_checked_at"],
            raw_score=row["raw_score"],
            final_score=row["final_score"],
            priority=row["priority"],
            scoring_version=row["scoring_version"],
            score_breakdown=row["score_breakdown_json"],
            scored_at=row["scored_at"],
            discovery_queries=row["discovery_queries"],
            google_maps_url=row["google_maps_url"],
        )


EXPORT_COLUMNS: tuple[str, ...] = (
    "business_id",
    "place_id",
    "business_name",
    "phone",
    "address",
    "rating",
    "review_count",
    "business_status",
    "original_website",
    "normalized_website",
    "final_website",
    "website_status",
    "http_status",
    "website_checked_at",
    "raw_score",
    "final_score",
    "priority",
    "scoring_version",
    "score_breakdown",
    "scored_at",
    "discovery_queries",
    "google_maps_url",
)

_NAVY = "FF17365D"
_WHITE = "FFFFFFFF"
_THIN_GRAY = "FFD9E2F3"
_PRIORITY_FILLS = {
    "very_high": "FFF4CCCC",
    "high": "FFFCE5CD",
    "medium": "FFFFF2CC",
    "low": "FFD9EAD3",
    "very_low": "FFD9E2F3",
}
_WEBSITE_STATUS_FILLS = {
    "working": "FFD9EAD3",
    "no_website": "FFF4CCCC",
    "unreachable": "FFFCE5CD",
    "dns_error": "FFFCE5CD",
    "ssl_error": "FFFCE5CD",
    "timeout": "FFFFF2CC",
    "http_error": "FFF4CCCC",
    "invalid_url": "FFF4CCCC",
    "blocked": "FFFFF2CC",
    "redirect_loop": "FFFCE5CD",
    "unknown_error": "FFEAD1DC",
}
_NUMBER_FORMATS = {
    "rating": "0.0",
    "review_count": "#,##0",
    "raw_score": "0",
    "final_score": "0",
    "http_status": "0",
}
_URL_COLUMNS = {
    "original_website",
    "normalized_website",
    "final_website",
    "google_maps_url",
}
_COLUMN_MAX_WIDTHS = {
    "business_id": 14,
    "place_id": 28,
    "business_name": 36,
    "phone": 20,
    "address": 44,
    "business_status": 22,
    "original_website": 48,
    "normalized_website": 48,
    "final_website": 48,
    "website_status": 20,
    "website_checked_at": 24,
    "priority": 16,
    "scoring_version": 18,
    "score_breakdown": 40,
    "scored_at": 24,
    "discovery_queries": 36,
    "google_maps_url": 48,
}


def sanitize_spreadsheet_text(value: str | None) -> str:
    """Prevent formula injection in CSV/XLSX by escaping dangerous characters."""
    if value is None:
        return ""
    text = str(value)
    stripped = text.lstrip()
    if stripped.startswith(("=", "+", "-", "@")):
        return f"'{text}"
    return text


def format_breakdown(json_str: str | None) -> str | None:
    """Format breakdown JSON compactly, returning None if invalid."""
    if not json_str:
        return None
    try:
        data = json.loads(json_str)
        return json.dumps(data, separators=(",", ":"))
    except json.JSONDecodeError:
        return None


def convert_row_for_export(row: LeadExportRow) -> dict[str, Any]:
    """Convert a row to ordered, formatted, spreadsheet-safe values."""
    result: dict[str, Any] = {}
    for column in EXPORT_COLUMNS:
        value = getattr(row, column)
        if column == "score_breakdown":
            value = format_breakdown(value)
        if isinstance(value, str):
            value = sanitize_spreadsheet_text(value)
        result[column] = value
    return result


def export_to_csv(rows: list[LeadExportRow], output_path: Path) -> None:
    """Export rows to a CSV file atomically."""
    fd, temp_path_str = tempfile.mkstemp(dir=output_path.parent, prefix="export_", suffix=".csv")
    temp_path = Path(temp_path_str)
    try:
        with open(fd, "w", encoding="utf-8", newline="") as handle:
            writer = csv.DictWriter(handle, fieldnames=EXPORT_COLUMNS)
            writer.writeheader()
            for row in rows:
                writer.writerow(convert_row_for_export(row))
        temp_path.replace(output_path)
    except Exception:
        temp_path.unlink(missing_ok=True)
        raise


def _column_positions() -> dict[str, int]:
    return {name: index for index, name in enumerate(EXPORT_COLUMNS, start=1)}


def _style_header(worksheet: Worksheet) -> None:
    fill = PatternFill(fill_type="solid", fgColor=_NAVY)
    border = Border(bottom=Side(style="thin", color=_WHITE))
    for cell in worksheet[1]:
        cell.fill = fill
        cell.font = Font(color=_WHITE, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    worksheet.row_dimensions[1].height = 30


def _style_body(worksheet: Worksheet) -> None:
    border = Border(
        left=Side(style="thin", color=_THIN_GRAY),
        right=Side(style="thin", color=_THIN_GRAY),
        top=Side(style="thin", color=_THIN_GRAY),
        bottom=Side(style="thin", color=_THIN_GRAY),
    )
    alignment = Alignment(vertical="top", wrap_text=True)
    for row_number, row in enumerate(worksheet.iter_rows(min_row=2), start=2):
        worksheet.row_dimensions[row_number].height = 31
        for cell in row:
            cell.border = border
            cell.alignment = alignment


def _apply_semantic_fills(worksheet: Worksheet) -> None:
    positions = _column_positions()
    priority_column = positions["priority"]
    website_column = positions["website_status"]
    for row_number in range(2, worksheet.max_row + 1):
        priority_cell = worksheet.cell(row=row_number, column=priority_column)
        priority_fill = _PRIORITY_FILLS.get(str(priority_cell.value))
        if priority_fill is not None:
            priority_cell.fill = PatternFill(fill_type="solid", fgColor=priority_fill)
            priority_cell.font = Font(bold=True)

        website_cell = worksheet.cell(row=row_number, column=website_column)
        website_fill = _WEBSITE_STATUS_FILLS.get(str(website_cell.value))
        if website_fill is not None:
            website_cell.fill = PatternFill(fill_type="solid", fgColor=website_fill)


def _apply_number_formats(worksheet: Worksheet) -> None:
    positions = _column_positions()
    for column, number_format in _NUMBER_FORMATS.items():
        column_index = positions[column]
        for row_number in range(2, worksheet.max_row + 1):
            worksheet.cell(row=row_number, column=column_index).number_format = number_format


def _make_urls_clickable(worksheet: Worksheet) -> None:
    positions = _column_positions()
    for column in _URL_COLUMNS:
        column_index = positions[column]
        for row_number in range(2, worksheet.max_row + 1):
            cell = cast(Cell, worksheet.cell(row=row_number, column=column_index))
            value = cell.value
            if isinstance(value, str) and value.startswith(("http://", "https://")):
                cell.hyperlink = value
                cell.font = Font(color="FF0563C1", underline="single")


def _set_content_aware_widths(worksheet: Worksheet) -> None:
    for column_index, column_name in enumerate(EXPORT_COLUMNS, start=1):
        maximum_content = max(
            (
                max((len(line) for line in str(cell.value).splitlines()), default=0)
                if cell.value is not None
                else 0
            )
            for cell in worksheet.iter_cols(
                min_col=column_index, max_col=column_index, min_row=1
            ).__next__()
        )
        maximum_width = _COLUMN_MAX_WIDTHS.get(column_name, 18)
        width = min(max(maximum_content + 2, 10), maximum_width)
        worksheet.column_dimensions[
            cast(Cell, worksheet.cell(row=1, column=column_index)).column_letter
        ].width = width


def _add_excel_table(worksheet: Worksheet) -> None:
    if worksheet.max_row < 2:
        return
    end_cell = worksheet.cell(row=worksheet.max_row, column=len(EXPORT_COLUMNS))
    start_cell = worksheet.cell(row=1, column=1)
    table = Table(
        displayName="LeadFinderLeads", ref=f"{start_cell.coordinate}:{end_cell.coordinate}"
    )
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    worksheet.add_table(table)
    worksheet.auto_filter.ref = table.ref


def _hide_technical_columns(worksheet: Worksheet) -> None:
    positions = _column_positions()
    for column in (
        "business_id",
        "place_id",
        "normalized_website",
        "scoring_version",
        "score_breakdown",
    ):
        cell = cast(Cell, worksheet.cell(row=1, column=positions[column]))
        worksheet.column_dimensions[cell.column_letter].hidden = True


def _configure_sheet_view_and_printing(worksheet: Worksheet) -> None:
    positions = _column_positions()
    freeze_column = positions["business_name"] + 1
    worksheet.freeze_panes = worksheet.cell(row=2, column=freeze_column).coordinate
    worksheet.sheet_view.showGridLines = False
    worksheet.print_title_rows = "1:1"
    worksheet.page_setup.orientation = "landscape"
    worksheet.page_setup.fitToWidth = 1
    worksheet.page_setup.fitToHeight = 0
    worksheet.sheet_properties.pageSetUpPr = PageSetupProperties(
        fitToPage=True, autoPageBreaks=False
    )


def _build_xlsx_workbook(rows: list[LeadExportRow]) -> openpyxl.Workbook:
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    if worksheet is None:
        worksheet = workbook.create_sheet("Leads")
    worksheet.title = "Leads"
    worksheet.append(EXPORT_COLUMNS)
    for row in rows:
        exported = convert_row_for_export(row)
        worksheet.append(
            [exported[column] if exported[column] is not None else "" for column in EXPORT_COLUMNS]
        )

    _style_header(worksheet)
    _style_body(worksheet)
    _apply_semantic_fills(worksheet)
    _apply_number_formats(worksheet)
    _make_urls_clickable(worksheet)
    _set_content_aware_widths(worksheet)
    _hide_technical_columns(worksheet)
    _add_excel_table(worksheet)
    _configure_sheet_view_and_printing(worksheet)
    return workbook


def export_to_xlsx(rows: list[LeadExportRow], output_path: Path) -> None:
    """Export rows to a professionally formatted XLSX file atomically."""
    fd, temp_path_str = tempfile.mkstemp(dir=output_path.parent, prefix="export_", suffix=".xlsx")
    temp_path = Path(temp_path_str)
    os.close(fd)
    workbook: openpyxl.Workbook | None = None
    try:
        workbook = _build_xlsx_workbook(rows)
        workbook.save(temp_path)
        workbook.close()
        workbook = None
        temp_path.replace(output_path)
    except Exception:
        if workbook is not None:
            workbook.close()
        temp_path.unlink(missing_ok=True)
        raise
