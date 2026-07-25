import csv
from pathlib import Path
from unittest.mock import patch

import openpyxl
import pytest

from lead_finder.exporter import (
    EXPORT_COLUMNS,
    LeadExportRow,
    export_to_csv,
    export_to_xlsx,
    format_breakdown,
    sanitize_spreadsheet_text,
)


class ExportTestError(RuntimeError):
    pass


@pytest.fixture
def sample_rows() -> list[LeadExportRow]:
    return [
        LeadExportRow(
            business_id=1,
            place_id="P1",
            business_name='Test "Business", with\nnewlines and ❤️',
            phone="+123",
            address="123 Street",
            rating=4.5,
            review_count=100,
            business_status="OPERATIONAL",
            original_website="http://test.com",
            normalized_website="https://test.com",
            final_website="https://test.com",
            website_status="working",
            http_status=200,
            website_checked_at="2023-01-01T00:00:00Z",
            raw_score=90,
            final_score=90,
            priority="very_high",
            scoring_version="v1",
            score_breakdown='{"timeout": 20, "working": 0}',
            scored_at="2023-01-01T00:00:00Z",
            discovery_queries="electrician",
            google_maps_url="https://maps.google.com/?cid=1",
        ),
        LeadExportRow(
            business_id=2,
            place_id="P2",
            business_name="=CMD()",  # Formula injection
            phone=None,
            address="-formula",
            rating=None,
            review_count=None,
            business_status="@mention",
            original_website=None,
            normalized_website=None,
            final_website=None,
            website_status=None,
            http_status=None,
            website_checked_at=None,
            raw_score=None,
            final_score=None,
            priority=None,
            scoring_version=None,
            score_breakdown=None,
            scored_at=None,
            discovery_queries=None,
            google_maps_url=None,
        ),
    ]


def test_sanitize_spreadsheet_text() -> None:
    assert sanitize_spreadsheet_text(None) == ""
    assert sanitize_spreadsheet_text("Hello") == "Hello"
    assert sanitize_spreadsheet_text("=CMD()") == "'=CMD()"
    assert sanitize_spreadsheet_text("+123") == "'+123"
    assert sanitize_spreadsheet_text("-abc") == "'-abc"
    assert sanitize_spreadsheet_text("@ref") == "'@ref"
    assert sanitize_spreadsheet_text(" =CMD()") == "' =CMD()"


def test_format_breakdown() -> None:
    assert format_breakdown(None) is None
    assert format_breakdown("invalid json") is None
    assert format_breakdown('{"a": 1, "b": 2}') == '{"a":1,"b":2}'


def test_export_to_csv_success(tmp_path: Path, sample_rows: list[LeadExportRow]) -> None:
    output = tmp_path / "test.csv"
    export_to_csv(sample_rows, output)

    assert output.exists()
    with open(output, encoding="utf-8", newline="") as f:
        reader = csv.DictReader(f)
        assert reader.fieldnames == list(EXPORT_COLUMNS)
        rows = list(reader)

        assert len(rows) == 2

        # UTF-8 preservation & CSV escaping for commas/quotes/newlines
        assert rows[0]["business_name"] == 'Test "Business", with\nnewlines and ❤️'

        # Format breakdown is compact
        assert rows[0]["score_breakdown"] == '{"timeout":20,"working":0}'

        # Empty instead of None
        assert rows[1]["phone"] == ""

        # Formula protection in CSV
        assert rows[1]["business_name"] == "'=CMD()"
        assert rows[1]["address"] == "'-formula"
        assert rows[1]["business_status"] == "'@mention"
        assert rows[0]["phone"] == "'+123"


def test_export_to_csv_parent_creation(tmp_path: Path, sample_rows: list[LeadExportRow]) -> None:
    output = tmp_path / "deep" / "dir" / "test.csv"
    output.parent.mkdir(parents=True)
    export_to_csv(sample_rows, output)
    assert output.exists()


def test_export_to_csv_atomic_failure(tmp_path: Path, sample_rows: list[LeadExportRow]) -> None:
    output = tmp_path / "test.csv"
    output.write_text("existing")

    with (
        patch(
            "csv.DictWriter.writerow",
            side_effect=ExportTestError("Failed"),
        ),
        pytest.raises(ExportTestError),
    ):
        export_to_csv(sample_rows, output)

    # Should remain unchanged
    assert output.read_text() == "existing"

    # Should not leave temporary files starting with export_ in the dir
    temp_files = list(output.parent.glob("export_*"))
    assert len(temp_files) == 0


def test_export_to_xlsx_success(tmp_path: Path, sample_rows: list[LeadExportRow]) -> None:
    output = tmp_path / "test.xlsx"
    export_to_xlsx(sample_rows, output)

    assert output.exists()
    wb = openpyxl.load_workbook(output)
    ws = wb.active
    assert ws.title == "Leads"

    # Check headers
    headers = [cell.value for cell in ws[1]]
    assert headers == list(EXPORT_COLUMNS)
    assert ws["A1"].font.bold is True

    # Freeze panes
    assert ws.freeze_panes == "D2"

    # Autofilter enabled
    assert ws.auto_filter.ref == "A1:V3"

    # Check values and types
    assert ws["C2"].value == 'Test "Business", with\nnewlines and ❤️'  # UTF-8
    assert ws["C3"].value == "'=CMD()"  # Formula injection
    assert ws["H3"].value == "'@mention"

    # Numeric types
    assert isinstance(ws["A2"].value, int)  # business_id
    assert ws["A2"].value == 1
    assert isinstance(ws["F2"].value, float)  # rating
    assert ws["F2"].value == 4.5
    assert isinstance(ws["M2"].value, int)  # http_status
    assert ws["M2"].value == 200

    positions = {name: index for index, name in enumerate(EXPORT_COLUMNS, start=1)}
    header = ws.cell(row=1, column=positions["business_name"])
    assert header.fill.fgColor.rgb == "FF17365D"
    assert header.font.color is not None
    assert header.font.color.rgb == "FFFFFFFF"
    assert header.font.bold is True
    assert ws.sheet_view.showGridLines is False
    assert ws.freeze_panes == "D2"
    assert ws.print_title_rows == "$1:$1"
    assert ws.page_setup.orientation == "landscape"

    assert "LeadFinderLeads" in ws.tables
    table = ws.tables["LeadFinderLeads"]
    assert table.tableStyleInfo is not None
    assert table.tableStyleInfo.showRowStripes is True

    priority_cell = ws.cell(row=2, column=positions["priority"])
    website_cell = ws.cell(row=2, column=positions["website_status"])
    assert priority_cell.fill.fill_type == "solid"
    assert website_cell.fill.fill_type == "solid"
    assert priority_cell.fill.fgColor.rgb != website_cell.fill.fgColor.rgb
    assert priority_cell.alignment.vertical == "top"
    assert priority_cell.alignment.wrap_text is True
    assert priority_cell.border.left.style == "thin"

    expected_formats = {
        "rating": "0.0",
        "review_count": "#,##0",
        "raw_score": "0",
        "final_score": "0",
        "http_status": "0",
    }
    for column, expected_format in expected_formats.items():
        assert ws.cell(row=2, column=positions[column]).number_format == expected_format

    for column in (
        "original_website",
        "normalized_website",
        "final_website",
        "google_maps_url",
    ):
        cell = ws.cell(row=2, column=positions[column])
        assert cell.hyperlink is not None
        assert cell.hyperlink.target == cell.value

    business_name_letter = ws.cell(row=1, column=positions["business_name"]).column_letter
    business_name_width = ws.column_dimensions[business_name_letter].width
    assert business_name_width is not None
    assert 15 < business_name_width <= 36

    technical_columns = (
        "business_id",
        "place_id",
        "normalized_website",
        "scoring_version",
        "score_breakdown",
    )
    for column in technical_columns:
        letter = ws.cell(row=1, column=positions[column]).column_letter
        assert ws.column_dimensions[letter].hidden is True

    breakdown_letter = ws.cell(row=1, column=positions["score_breakdown"]).column_letter
    breakdown_width = ws.column_dimensions[breakdown_letter].width
    assert breakdown_width is not None
    assert breakdown_width <= 40

    assert ws.cell(row=2, column=positions["business_id"]).value == 1
    assert ws.cell(row=2, column=positions["place_id"]).value == "P1"
    assert ws.cell(row=2, column=positions["normalized_website"]).value == "https://test.com"
    assert ws.cell(row=2, column=positions["scoring_version"]).value == "v1"
    assert ws.cell(row=2, column=positions["score_breakdown"]).value == (
        '{"timeout":20,"working":0}'
    )
    assert ws.row_dimensions[2].height == 31

    wb.close()


def test_export_to_xlsx_atomic_failure(tmp_path: Path, sample_rows: list[LeadExportRow]) -> None:
    output = tmp_path / "test.xlsx"
    output.write_bytes(b"existing")

    with (
        patch(
            "openpyxl.worksheet.worksheet.Worksheet.append",
            side_effect=ExportTestError("Failed"),
        ),
        pytest.raises(ExportTestError),
    ):
        export_to_xlsx(sample_rows, output)

    # Should remain unchanged
    assert output.read_bytes() == b"existing"

    # Should not leave temporary files starting with export_ in the dir
    temp_files = list(output.parent.glob("export_*"))
    assert len(temp_files) == 0
