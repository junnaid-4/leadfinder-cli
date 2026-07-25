"""Failure-safety and real pipeline integration tests."""

from __future__ import annotations

import csv
import sqlite3
from pathlib import Path

import httpx
import openpyxl
import pytest
import respx
import yaml
from typer.testing import CliRunner

from lead_finder.cli import app
from lead_finder.config import AppConfig
from lead_finder.database import Database, init_database
from lead_finder.pipeline import (
    WebsiteCheckStageError,
    check_business_websites,
    collect_businesses,
    export_lead_files,
)
from lead_finder.places_client import PLACES_API_URL

runner = CliRunner()
REPOSITORY_ROOT = Path(__file__).resolve().parents[1]


def _demo_config(tmp_path: Path) -> Path:
    raw = yaml.safe_load((REPOSITORY_ROOT / "config.demo.yaml").read_text(encoding="utf-8"))
    raw["database"]["path"] = str(tmp_path / "data" / "demo_lead_finder.db")
    raw["export"]["output_directory"] = str(tmp_path / "demo_output")
    raw["output"]["directory"] = str(tmp_path / "demo_output")
    path = tmp_path / "config.demo.yaml"
    path.write_text(yaml.safe_dump(raw), encoding="utf-8")
    return path


def test_demo_refuses_normal_and_arbitrary_targets(tmp_path: Path) -> None:
    protected_db = tmp_path / "arbitrary.db"
    protected_output = tmp_path / "arbitrary-output"
    protected_db.write_bytes(b"protected database")
    protected_output.mkdir()
    protected_csv = protected_output / "leads.csv"
    protected_csv.write_bytes(b"protected csv")
    raw = yaml.safe_load((REPOSITORY_ROOT / "config.demo.yaml").read_text(encoding="utf-8"))
    raw["demo"]["enabled"] = False
    raw["database"]["path"] = str(protected_db)
    raw["export"]["output_directory"] = str(protected_output)
    path = tmp_path / "normal.yaml"
    path.write_text(yaml.safe_dump(raw), encoding="utf-8")

    result = runner.invoke(app, ["demo", "--config", str(path), "--overwrite"])
    assert result.exit_code == 1
    assert protected_db.read_bytes() == b"protected database"
    assert protected_csv.read_bytes() == b"protected csv"

    raw["demo"]["enabled"] = True
    path.write_text(yaml.safe_dump(raw), encoding="utf-8")
    result = runner.invoke(app, ["demo", "--config", str(path), "--overwrite"])
    assert result.exit_code == 1
    assert protected_db.read_bytes() == b"protected database"
    assert protected_csv.read_bytes() == b"protected csv"

    project = tmp_path / "project"
    project.mkdir()
    outside_database = tmp_path / "outside" / "data" / "demo_lead_finder.db"
    outside_output = tmp_path / "outside" / "demo_output"
    outside_database.parent.mkdir(parents=True)
    outside_output.mkdir(parents=True)
    outside_database.write_bytes(b"outside database")
    (outside_output / "leads.csv").write_bytes(b"outside csv")
    raw["database"]["path"] = str(outside_database)
    raw["export"]["output_directory"] = str(outside_output)
    outside_config = project / "config.demo.yaml"
    outside_config.write_text(yaml.safe_dump(raw), encoding="utf-8")
    result = runner.invoke(app, ["demo", "--config", str(outside_config), "--overwrite"])
    assert result.exit_code == 1
    assert outside_database.read_bytes() == b"outside database"
    assert (outside_output / "leads.csv").read_bytes() == b"outside csv"


def test_demo_xlsx_failure_preserves_existing_artifacts(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    config = _demo_config(tmp_path)
    database = tmp_path / "data" / "demo_lead_finder.db"
    output = tmp_path / "demo_output"
    database.parent.mkdir()
    output.mkdir()
    database.write_bytes(b"old database")
    (output / "leads.csv").write_bytes(b"old csv")
    (output / "leads.xlsx").write_bytes(b"old xlsx")

    def fail_xlsx(*args: object, **kwargs: object) -> None:
        raise RuntimeError("injected XLSX failure")

    monkeypatch.setattr("lead_finder.pipeline.export_to_xlsx", fail_xlsx)
    result = runner.invoke(app, ["demo", "--config", str(config), "--overwrite"])
    assert result.exit_code == 1
    assert "injected XLSX failure" in result.stdout
    assert database.read_bytes() == b"old database"
    assert (output / "leads.csv").read_bytes() == b"old csv"
    assert (output / "leads.xlsx").read_bytes() == b"old xlsx"
    assert not list(tmp_path.glob(".leadfinder-demo-*"))


def _scored_config(tmp_path: Path) -> AppConfig:
    config = AppConfig.model_validate(
        {
            "project": {"name": "Export Safety"},
            "search": {"location": "Test", "queries": ["test"]},
            "database": {"path": str(tmp_path / "test.db")},
            "export": {"output_directory": str(tmp_path / "exports")},
        }
    )
    db = init_database(config.database_path())
    db.execute(
        "INSERT INTO businesses (id, place_id, business_name, business_status) "
        "VALUES (1, 'p1', 'Fictional One', 'OPERATIONAL')"
    )
    db.save_lead_score(1, 50, 50, "medium", "[]", "v1", "2026-01-01T00:00:00Z")
    db.close()
    return config


def test_both_export_failure_changes_neither_destination(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    config = _scored_config(tmp_path)
    output = config.export_directory()
    output.mkdir()
    csv_path = output / "leads.csv"
    xlsx_path = output / "leads.xlsx"
    csv_path.write_bytes(b"old csv")
    xlsx_path.write_bytes(b"old xlsx")

    def fail_xlsx(*args: object, **kwargs: object) -> None:
        raise RuntimeError("injected XLSX failure")

    monkeypatch.setattr("lead_finder.pipeline.export_to_xlsx", fail_xlsx)
    with pytest.raises(RuntimeError, match="injected XLSX failure"):
        export_lead_files(config, fmt="both", overwrite=True)
    assert csv_path.read_bytes() == b"old csv"
    assert xlsx_path.read_bytes() == b"old xlsx"


def test_website_persistence_failure_is_stage_failure(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    config = AppConfig.model_validate(
        {
            "project": {"name": "Website Failure"},
            "search": {"location": "Test", "queries": ["test"]},
            "database": {"path": str(tmp_path / "test.db")},
        }
    )
    db = init_database(config.database_path())
    db.execute(
        "INSERT INTO businesses (id, place_id, business_name, website_url) "
        "VALUES (1, 'p1', 'Fictional One', NULL)"
    )
    db.commit()
    db.close()

    def fail_save(self: Database, *args: object, **kwargs: object) -> None:
        raise sqlite3.OperationalError("disk unavailable")

    monkeypatch.setattr(Database, "save_website_check_result", fail_save)
    with pytest.raises(WebsiteCheckStageError):
        check_business_websites(config)
    config_path = tmp_path / "config.yaml"
    config_path.write_text(yaml.safe_dump(config.model_dump(mode="json")), encoding="utf-8")
    cli_result = runner.invoke(app, ["check-websites", "--config", str(config_path)])
    assert cli_result.exit_code == 1
    assert "Website checking failed" in cli_result.stdout
    connection = sqlite3.connect(config.database_path())
    try:
        assert connection.execute("SELECT count(*) FROM website_checks").fetchone()[0] == 0
    finally:
        connection.close()


def test_collection_unexpected_failure_marks_run_failed(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    config = AppConfig.model_validate(
        {
            "project": {"name": "Collection Failure"},
            "search": {"location": "Test", "queries": ["test"]},
            "database": {"path": str(tmp_path / "test.db")},
        }
    )

    async def fail_query(*args: object, **kwargs: object) -> int:
        raise RuntimeError("injected query logging failure")

    monkeypatch.setattr("lead_finder.pipeline._collect_query", fail_query)
    with pytest.raises(RuntimeError, match="injected query logging failure"):
        collect_businesses(config, "test-key")
    connection = sqlite3.connect(config.database_path())
    try:
        status = connection.execute("SELECT status FROM search_runs").fetchone()[0]
    finally:
        connection.close()
    assert status == "FAILED"


@respx.mock
def test_real_full_pipeline_handoff_and_exports(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    config_path = tmp_path / "config.yaml"
    config_path.write_text(
        yaml.safe_dump(
            {
                "project": {"name": "Integration"},
                "search": {
                    "location": "Exampleton, Fictionland",
                    "queries": ["fictional services"],
                    "max_results_per_query": 2,
                    "max_total_results": 2,
                    "max_api_requests": 2,
                },
                "database": {"path": str(tmp_path / "pipeline.db")},
                "logging": {
                    "file": False,
                    "console": False,
                    "directory": str(tmp_path / "logs"),
                },
                "export": {"output_directory": str(tmp_path / "exports")},
            }
        ),
        encoding="utf-8",
    )
    monkeypatch.setenv("GOOGLE_MAPS_API_KEY", "test-key")
    places_route = respx.post(PLACES_API_URL).mock(
        return_value=httpx.Response(
            200,
            json={
                "places": [
                    {
                        "id": "demo-integration-1",
                        "displayName": {"text": "Fictional High Lead"},
                        "formattedAddress": "1 Demo Way, Fictionland",
                        "nationalPhoneNumber": "+1 202-555-0151",
                        "rating": 4.9,
                        "userRatingCount": 150,
                        "businessStatus": "OPERATIONAL",
                    },
                    {
                        "id": "demo-integration-2",
                        "displayName": {"text": "Fictional Low Lead"},
                        "formattedAddress": "2 Demo Way, Fictionland",
                        "websiteUri": "https://integration.example.com",
                        "rating": 2.5,
                        "userRatingCount": 1,
                        "businessStatus": "OPERATIONAL",
                    },
                ]
            },
        )
    )
    website_route = respx.get("https://integration.example.com").mock(
        return_value=httpx.Response(200, headers={"content-type": "text/html"})
    )

    result = runner.invoke(
        app,
        ["run", "--config", str(config_path), "--format", "both"],
    )
    assert result.exit_code == 0, result.stdout
    assert places_route.call_count == 1
    assert website_route.call_count == 1

    connection = sqlite3.connect(tmp_path / "pipeline.db")
    try:
        assert connection.execute("SELECT count(*) FROM businesses").fetchone()[0] == 2
        assert connection.execute("SELECT count(*) FROM website_checks").fetchone()[0] == 2
        assert connection.execute("SELECT count(*) FROM lead_scores").fetchone()[0] == 2
        priorities = {
            row[0] for row in connection.execute("SELECT DISTINCT priority FROM lead_scores")
        }
    finally:
        connection.close()
    assert len(priorities) >= 2

    csv_path = tmp_path / "exports" / "leads.csv"
    with csv_path.open(encoding="utf-8", newline="") as handle:
        csv_rows = list(csv.DictReader(handle))
    assert {row["business_name"] for row in csv_rows} == {
        "Fictional High Lead",
        "Fictional Low Lead",
    }

    workbook = openpyxl.load_workbook(tmp_path / "exports" / "leads.xlsx")
    try:
        values = {
            str(workbook.active.cell(row=index, column=3).value)
            for index in range(2, workbook.active.max_row + 1)
        }
    finally:
        workbook.close()
    assert values == {"Fictional High Lead", "Fictional Low Lead"}
